#!/usr/bin/env python3
"""
Regenerates public/data/tours-index.json and public/data/details/*.json
from the master Europe Tours Excel file.

Run this whenever:
  - The yearly full reprice update lands in the master Excel file
  - A batch of new Tour Highlights is added
  - Any bulk change happens upstream that should flow to the live site

Usage:
    python3 generate_site_data.py /path/to/Classic_Vacations_Preferred_Europe_Tours.xlsx

After running this, commit and push the changed files in public/data/ —
Netlify will auto-redeploy from the new commit, same as a staff one-off edit.

IMPORTANT: this OVERWRITES tours-index.json and all files in details/.
Any one-off staff edits made via the admin panel since the last regeneration
will be replaced by whatever the master Excel file says. That's intentional —
the master file is the single source of truth once a real bulk update lands;
one-off edits exist only to patch gaps between bulk updates.
"""
import sys
import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

BLOCK_HEADERS = {"STEP by STEP:", "BASIC INFO:", "INCLUSIONS and RESTRICTIONS:",
                  "NOT INCLUDED:", "IMPORTANT GENERAL NOTES:"}

# Auto-detected theme tags — layered on top of the supplier-authored "Tour Type
# Tags" column. These exist because "Food/Wine" and "Culture/History" are each
# broad enough (1,360 and 386 tours respectively) that filtering by category
# alone doesn't actually narrow anything down.
THEME_PATTERNS = [
    ("Wine Tasting/Vineyard", re.compile(r"\b(wine tasting|vineyard|winery|wine estate|wine experience)\b", re.I)),
    ("Cooking Class", re.compile(r"cooking class|cooking experience|hands-on cook|culinary class", re.I)),
    ("Castle/Palace", re.compile(r"\b(castle|palace|chateau|schloss)\b", re.I)),
    ("Museum/Art", re.compile(r"\b(museum|gallery|masterpiece)\b", re.I)),
    ("Church/Religious", re.compile(r"\b(church|cathedral|basilica|synagogue|monastery|abbey)\b", re.I)),
]


def detect_themes(*texts):
    haystack = " ".join(t for t in texts if isinstance(t, str))
    return [name for name, pattern in THEME_PATTERNS if pattern.search(haystack)]


def is_boundary(line):
    s = line.strip()
    return s in BLOCK_HEADERS or s.startswith("Welcome Experience includes:")


def dedupe_details(text):
    if not isinstance(text, str):
        return ""
    lines = text.split('\n')
    chunks, current = [], []
    for line in lines:
        if is_boundary(line):
            if current:
                chunks.append(current)
            current = [line]
        else:
            current.append(line)
    if current:
        chunks.append(current)
    seen, out = set(), []
    for c in chunks:
        key = tuple(l.strip() for l in c)
        if key in seen:
            continue
        seen.add(key)
        out.append(c)
    return '\n\n'.join('\n'.join(c).strip('\n') for c in out if any(l.strip() for l in c))


def parse_highlights(text, content, tour):
    if isinstance(text, str) and text.strip():
        lines = [l for l in text.split('\n') if l.strip()]
        intro = lines[0].strip()
        bullets = [l.strip().lstrip('-').strip() for l in lines[1:] if l.strip().startswith('-')]
        return intro, bullets, True
    c = str(content) if isinstance(content, str) else ""
    if c.startswith(tour):
        c = c[len(tour):].strip(' -')
    sentences = re.split(r'(?<=[.!?])\s+', c)
    intro = ' '.join(sentences[:2]).strip()[:280]
    return intro, [], False


def clean_content(content, tour):
    c = str(content) if isinstance(content, str) else ""
    if not c.strip() or c.strip().lower() == "nan":
        return ""
    if c.startswith(tour):
        c = c[len(tour):].strip(' -')
    return c.strip()


def should_include_overview(content_text, intro):
    if not content_text:
        return False
    c_norm = ' '.join(content_text.split())
    i_norm = ' '.join(intro.split())
    if not c_norm or c_norm == i_norm:
        return False
    return len(c_norm) > len(i_norm) + 20


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 generate_site_data.py /path/to/master.xlsx")
        sys.exit(1)

    src_path = Path(sys.argv[1])
    out_dir = Path(__file__).parent.parent / "public" / "data"
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(src_path)
    ws = wb["Europe Tours"]
    hidden_rows = set(r for r in range(22, ws.max_row + 1) if ws.row_dimensions[r].hidden)

    df = pd.read_excel(src_path, sheet_name="Europe Tours", header=20)
    df["excel_row"] = df.index + 22
    df = df[~df["excel_row"].isin(hidden_rows)].reset_index(drop=True)

    print(f"Total rows in master file: {len(df) + len(hidden_rows)}")
    print(f"Hidden (structurally excluded, 3+ pax only): {len(hidden_rows)}")
    print(f"Visible rows going to the site: {len(df)}")

    index_items = []
    details_map = {}
    for _, r in df.iterrows():
        price_raw = r["Starting Price for 2 pax"]
        is_tbd = pd.isna(price_raw) or str(price_raw).strip().upper() == "TBD"
        price_display = "Rate available in Plex" if is_tbd else (
            str(price_raw) if str(price_raw).startswith("$") else f"${price_raw}"
        )

        tags = str(r["Tour Type Tags"]) if pd.notna(r["Tour Type Tags"]) else ""
        tags_list = [t.strip() for t in tags.split(";") if t.strip()][:5]

        intro, bullets, ready = parse_highlights(r["Tour Highlights"], r["Content"], str(r["Tour"]))
        tour_id = f"row{int(r['excel_row'])}"
        themes = detect_themes(str(r["Tour"]), r["Content"], r["Tour Highlights"], r["Details"])

        index_items.append({
            "id": tour_id,
            "tour": str(r["Tour"]),
            "country": r["Country"] if pd.notna(r["Country"]) else "",
            "city": r["City"] if pd.notna(r["City"]) else "",
            "category": r["Tour Category"] if pd.notna(r["Tour Category"]) else "",
            "style": r["Service Style"] if pd.notna(r["Service Style"]) else "",
            "duration": r["Duration"] if pd.notna(r["Duration"]) else "",
            "priceDisplay": price_display,
            "priceIsPlex": is_tbd,
            "summaryIntro": intro,
            "summaryBullets": bullets,
            "highlightsReady": ready,
            "tags": tags_list,
            "themes": themes,
        })
        content_text = clean_content(r["Content"], str(r["Tour"]))
        details_text = dedupe_details(str(r["Details"]) if pd.notna(r["Details"]) else "")
        parts = []
        if should_include_overview(content_text, intro):
            parts.append("OVERVIEW:\n" + content_text)
        if details_text:
            parts.append(details_text)
        details_map[tour_id] = "\n\n".join(parts)

    with open(out_dir / "tours-index.json", "w") as f:
        json.dump({"generatedAt": pd.Timestamp.now().strftime("%Y-%m-%d"),
                    "count": len(index_items), "tours": index_items}, f)

    with open(out_dir / "tours-details.json", "w") as f:
        json.dump(details_map, f)

    print(f"Wrote {len(index_items)} tours to tours-index.json and tours-details.json")
    print("Next: commit and push public/data/ to GitHub — GitHub Pages redeploys automatically.")


if __name__ == "__main__":
    main()
